from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import os
import csv
from db import lift, pour

ROOT_FOLDER = os.getcwd()
CRA_FOLDER = Path(ROOT_FOLDER) / "CRA"

def save_csv(extracted_data:list[dict], headers: list[str], new_filename: str) -> None:
    """Save the data extracted from excel file (xlsx) in csv file

    Args:
        extracted_data (list[dict]): List of dictionaries with the data
                                     extracted from each row of the excel file.
        headers (list[str]): List of fields (column names) and keys of the dicts in extracted_data.
        new_filename (str): Name of the csv file that will be generated.
    """
    with open(new_filename, 'w',newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(extracted_data)
        print(f"File {new_filename} generated successfully...")


def read_csv(path_file: str,) -> list[dict]:
    with open(path_file, 'r', encoding='utf-8') as csv_file:
        reader = csv.DictReader(csv_file)

        # Convert to dictionary
        result = [row for row in reader]
        return result
    


def main():
    data = []
     # Read csv file
    cra_rebar_pac_file = Path(ROOT_FOLDER) / "data_cra_rebar_not_found1.xlsx"
    wb = load_workbook(cra_rebar_pac_file)
    sheetname = wb['data_cra_rebar_not_found']
    start_row = 2
    end_row = 348
    for i in range(start_row, end_row):
        original_lift = sheetname.cell(row=i, column=1).value
        lift_name = sheetname.cell(row=i, column=2).value
        corrected = sheetname.cell(row=i, column=3).value
        rebar_units = sheetname.cell(row=i, column=4).value
        info = {}
        info['original_lift'] = original_lift
        info['lift'] = lift_name
        info['corrected'] = corrected
        info['rebar_units'] = rebar_units

        # Get the actual_start date
        lift_id = lift.get_lift_id_by_name(corrected)
        actual_start = pour.get_lift_actual_start(lift_id)
        info['actual_start'] = actual_start
        data.append(info)


    headers = ["original_lift", "lift", "corrected", "rebar_units", "actual_start"]
    save_csv(data, headers, "lifts_not_found_with_actual_start.csv")

    
    for item in data:
        print(item)




if __name__ == '__main__':
    main()